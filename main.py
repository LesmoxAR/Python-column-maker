import os
from openpyxl import load_workbook
from openpyxl.styles import Font



def convert_to_columns(file_path):
    # Carga el archivo CSV
    with open(file_path, 'r') as f:
        csv_data = f.read()
        
    # Crea un nuevo archivo Excel
    workbook = load_workbook()
    
    # Accede a la primera hoja
    worksheet = workbook.active
    
    # Convierte la información a una lista de filas y columnas
    rows = csv_data.split('\n')
    data = [r.split(',') for r in rows]
    
    # Encuentra los índices de las columnas de Full Name, First Seen y Time in Call
    full_name_column = None
    first_seen_column = None
    time_in_call_column = None
    for i, cell_value in enumerate(data[4]):
        if cell_value == "Full Name":
            full_name_column = i
        elif cell_value == "First Seen":
            first_seen_column = i
        elif cell_value == "Time in Call":
            time_in_call_column = i
    
    # Inserta una nueva columna a la derecha de la columna "Full Name"
    worksheet.insert_cols(full_name_column+1)
    # Agrega el encabezado de la nueva columna
    worksheet.cell(row=1, column=full_name_column+1, value="First Seen")
    # Obtiene la letra de la nueva columna
    first_seen_column_letter = worksheet.cell(row=1, column=full_name_column+1).column_letter

    # Inserta una nueva columna a la derecha de la columna "First Seen"
    worksheet.insert_cols(first_seen_column+2)
    # Agrega el encabezado de la nueva columna
    worksheet.cell(row=1, column=first_seen_column+2, value="Time in Call")
    # Obtiene la letra de la nueva columna
    time_in_call_column_letter = worksheet.cell(row=1, column=first_seen_column+2).column_letter
    
    # Itera sobre cada fila de datos
    for i, row_data in enumerate(data[5:]):
        if row_data:
            # Separa el texto de la celda en 3 partes
            full_name, first_seen, time_in_call = row_data[full_name_column], row_data[first_seen_column], row_data[time_in_call_column]
            # Ingresa los datos en las nuevas columnas
            worksheet.cell(row=i+6, column=full_name_column+1, value=first_seen)
            worksheet.cell(row=i+6, column=first_seen_column+2, value=time_in_call)
            # Ingresa el nombre completo en la columna original
            worksheet.cell(row=i+6, column=full_name_column, value=full_name)
    
    # Aplica el formato en negrita a las nuevas columnas
    worksheet.column_dimensions[first_seen_column_letter].width = 20
    worksheet.column_dimensions[time_in_call_column_letter].width = 15
    bold_font = Font(bold=True)
    worksheet.cell(row=1, column=full_name_column).font = bold_font
    worksheet.cell(row=1, column=full_name_column+1).font = bold_font
    worksheet.cell(row=1, column=full_name_column+2).font = bold_font
    
    # Guarda el archivo
    workbook.save(file_path.replace(".csv", ".xlsx"))
    print


file = input("Ingrese la ruta del archivo: ")
convert_to_columns(file)